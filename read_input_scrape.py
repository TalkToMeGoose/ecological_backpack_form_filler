'''

The purpose of this is to:
1) take values out of a csv file,
2) plug them into and online survey
link: https://www.ressourcen-rechner.de/calculator.php?lang=en
3) scrape the result
4) insert back into python
notes:
The waiting and next button commands, while inelegant,
are necessary to execute the javascript survey

'''
#first install selenium, pandas, and xlrd
import csv
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.support.ui import Select
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.remote.webelement import WebElement
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as ec
import os.path
import time
import pandas as pd
import openpyxl
import numpy as np

'''
WORKBOOK SETUP
Creates workbook. We iteratively write the footprint 
results to this as we scrape them from the Wuppertal survey.
These will be merged later in R to the derived data
'''

# if workbook exists, load it
if os.path.exists('impact_results.xlsx'):
    wb = openpyxl.load_workbook('impact_results.xlsx')
    sheet = wb.active
else:
    # else, create new workbook
    wb = openpyxl.Workbook()
    sheet = wb.active
    
    # and add titles in the first row of each column
    sheet.cell(row=1, column=1).value = 'CASE'
    sheet.cell(row=1, column=2).value = 'EI_HO'
    sheet.cell(row=1, column=3).value = 'EI_CS'
    sheet.cell(row=1, column=4).value = 'EI_NU'
    sheet.cell(row=1, column=5).value = 'EI_LE'
    sheet.cell(row=1, column=6).value = 'EI_MO'
    sheet.cell(row=1, column=7).value = 'EI_VA'
    sheet.cell(row=1, column=8).value = 'EI_SUM'

# read thesis survey results
dataframe = pd.read_excel('sosci_survey_derived.xlsx')

'''
DATA MANIPULATION SECTION
To make the data fit into the online survey
some of it must be manipulated (particularly
for the radio buttons)
'''

#dataframe['LANGUAGE'] = dataframe['LANGUAGE'].replace(['eng','ger'],['en','de'])
dataframe['HO04'] = dataframe['HO04'].replace([1,2,3,-1], [41,42,43,44])
dataframe['HO05'] = dataframe['HO05'].replace([1,2,3,4], [46,47,48,49])
dataframe['HO06'] = dataframe['HO06'].replace([1,2,3,4,5,6,7,-1], [51,52,53,54,55,56,57,58])
dataframe['HO07'] = dataframe['HO07'].replace([1,2,3,4], [60,61,62,63])
dataframe['CS07'] = dataframe['CS07'].replace([1,2], [87,88])
dataframe['NU01'] = dataframe['NU01'].replace([1,2,3,4], [90,91,92,93])
dataframe['NU03'] = dataframe['NU03'].replace([1,2,3,4], [106,107,108,109])
dataframe['LE03'] = dataframe['LE03'].replace([1,2], [128,129])
dataframe['MO04'] = dataframe['MO04'].replace([1,2,3,4,5,6,7,8,-1], [150,151,152,153,154,155,156,157,158])
dataframe['MO05'] = dataframe['MO05'].replace([1,2,3,4,5,6,7,-1], [160,161,162,163,164,165,166,167])
dataframe['MO06'] = dataframe['MO06'].replace([1,-1], [169,170])
dataframe['MO07'] = dataframe['MO07'].replace([1,2,3], [172,173,174])

# for debugging
entry = dataframe.iloc[0]

'''
BROWSER SETUP SECTION
ensure you have a version of chromedriver that matches your browser: https://chromedriver.chromium.org/downloads
'''
# set up pathways
s=Service('C:/Users/Beemo/chromedriver/chromedriver.exe')
brave_path = 'C:/Program Files/BraveSoftware/Brave-Browser/Application/brave.exe'

option = webdriver.ChromeOptions()
option.binary_location = brave_path

# opens blank browser window
browser = webdriver.Chrome(service=s, options = option)
# sets and gets url
url='https://www.ressourcen-rechner.de/calculator.php?lang=en/'
browser.get(url)

# make wait funciton with webdriver
wait = WebDriverWait(browser, 30)


'''
SURVEY SECTION
'''

# make next page function
def next_page():
    time.sleep(0.5)
    browser.find_element(By.CSS_SELECTOR, 'p[id="next-button"]').click()
    time.sleep(1)
    
def run_survey():
    # pause for debugging
    time.sleep(1)
    
    #---------------------------------
    # SECTION 1: Housing
    #---------------------------------
    # print case for console logging
    print(entry['CASE'])
    
    # select language
    lang_select = Select(browser.find_element(By.ID, 'lang-selector'))
    # chooses english to avoid "," / "." discrepancy
    lang_select.select_by_value('en')

    time.sleep(0.5)

    # HO01 is the intro paragraph
    # HO02 place of residence
    # Wuppetal survey indexes to 0 while SoSci indexes at 1; we subtract 1 from the entry below
    country_select = Select(browser.find_element(By.ID, 'answers01'))
    country_select.select_by_index(int(entry['HO02']-1))

    next_page()

    # HO03 (input): # Adults, # children, & size of Household
    adult_input = browser.find_element(By.ID, '37')
    adult_input.send_keys(entry['HO03_01'])

    child_input = browser.find_element(By.ID, '38')
    child_input.send_keys(entry['HO03_02'])

    house_size_input = browser.find_element(By.ID, '39')
    house_size_input.send_keys(entry['HO03_04'])

    next_page()

    # HO04 (radio): Electricity type
    elec_type_xpath = f"//input[@name=\"Q03\"][@id=\"{int(entry['HO04'])}\"]"
    elec_type_radio = browser.find_element(By.XPATH, elec_type_xpath)
    browser.execute_script("return arguments[0].click()", elec_type_radio)

    next_page()

    # HO05 (radio + input): Electricity usage
    elec_usage_xpath = f"//input[@name=\"Q04\"][@id=\"{int(entry['HO05'])}\"]"
    elec_usage_radio = browser.find_element(By.XPATH, elec_usage_xpath)
    browser.execute_script("return arguments[0].click()", elec_usage_radio)

    time.sleep(0.5)

    if entry['HO05_01'] > 0:
        elec_usage_input = browser.find_element(By.NAME, 'optValQ04')
        elec_usage_input.send_keys(entry['HO05_01'])

    next_page()

    # HO06 (radio): Heating type
    heat_type_xpath = f"//input[@name=\"Q05\"][@id=\"{int(entry['HO06'])}\"]"
    heat_type_radio = browser.find_element(By.XPATH, heat_type_xpath)
    browser.execute_script("return arguments[0].click()", heat_type_radio)

    next_page()

    # HO07 (radio + input): Heating usage
    heat_usage_xpath = f"//input[@name=\"Q06\"][@id=\"{int(entry['HO07'])}\"]"
    heat_usage_radio = browser.find_element(By.XPATH, heat_usage_xpath)
    browser.execute_script("return arguments[0].click()", heat_usage_radio)

    time.sleep(0.5)

    if entry['HO07_01'] > 0:
        heat_usage_input = browser.find_element(By.NAME, 'optValQ06')
        heat_usage_input.send_keys(entry['HO07_01'])

    next_page()
    # wait until the number appears
    wait.until(ec.visibility_of_element_located((By.XPATH, '//*[@id="resultArea"]/div[3]/h2/strong')))
    # and scrape it off the website
    # results saved belwo as attribute (e.g. run_survey.result_housing)
    # so that they can be called outside of the function
    run_survey.result_housing = browser.find_element(By.XPATH, '//*[@id="resultArea"]/div[3]/h2/strong').text
    print("Housing:", run_survey.result_housing, "tons")
    next_page()

    #---------------------------------
    # SECTION 2: Consumer Goods
    #---------------------------------

    # CS01: Household Appliances
    oven_input = browser.find_element(By.ID, '65')
    oven_input.send_keys(entry['CS01_01'])

    micro_input = browser.find_element(By.ID, '66')
    micro_input.send_keys(entry['CS01_02'])

    fridge_input = browser.find_element(By.ID, '67')
    fridge_input.send_keys(entry['CS01_03'])

    freezer_input = browser.find_element(By.ID, '68')
    freezer_input.send_keys(entry['CS01_04'])

    dish_input = browser.find_element(By.ID, '69')
    dish_input.send_keys(entry['CS01_05'])

    washer_input = browser.find_element(By.ID, '70')
    washer_input.send_keys(entry['CS01_06'])

    dryer_input = browser.find_element(By.ID, '71')
    dryer_input.send_keys(entry['CS01_07'])

    # CS02 check if appliances are used until broken
    if entry['CS02_01'] == 2:
        browser.find_element(By.ID, '72').click()

    next_page()

    # CS03: Consumer Electronics
    tv_input = browser.find_element(By.ID, '74')
    tv_input.send_keys(entry['CS03_01'])

    dvd_input = browser.find_element(By.ID, '75')
    dvd_input.send_keys(entry['CS03_02'])

    radio_input = browser.find_element(By.ID, '76')
    radio_input.send_keys(entry['CS03_03'])

    pc_input = browser.find_element(By.ID, '77')
    pc_input.send_keys(entry['CS03_04'])

    notebook_input = browser.find_element(By.ID, '78')
    notebook_input.send_keys(entry['CS03_05'])

    tablet_input = browser.find_element(By.ID, '79')
    tablet_input.send_keys(entry['CS03_06'])

    mobile_input = browser.find_element(By.ID, '80')
    mobile_input.send_keys(entry['CS03_07'])

    gaming_input = browser.find_element(By.ID, '81')
    gaming_input.send_keys(entry['CS03_08'])

    # CS04 check if appliances are used until broken
    if entry['CS04_01'] == 2:
        browser.find_element(By.ID, '82').click()

    next_page()

    # CS05 Clothing
    clothing_input = browser.find_element(By.ID, '84')
    clothing_input.send_keys(entry['CS05_01'])

    # CS06 check if clothing bought second hand
    if entry['CS06_01'] == 2:
        browser.find_element(By.ID, '85').click()

    next_page()

    # CS07 furniture new or old
    furniture_age_xpath = f"//input[@name=\"Q10\"][@id=\"{int(entry['CS07'])}\"]"
    furniture_age_radio = browser.find_element(By.XPATH, furniture_age_xpath)
    browser.execute_script("return arguments[0].click()", furniture_age_radio)

    next_page()
    wait.until(ec.visibility_of_element_located((By.XPATH, '//*[@id="resultArea"]/div[3]/h2/strong')))
    run_survey.result_consumer = browser.find_element(By.XPATH, '//*[@id="resultArea"]/div[3]/h2/strong').text
    print("Consumer goods:", run_survey.result_consumer, "tons")
    next_page()

    #---------------------------------
    # SECTION 3: Nutrition
    #---------------------------------

    # NU01: Diet (meat intake)
    meat_intake_xpath = f"//input[@name=\"Q11\"][@id=\"{int(entry['NU01'])}\"]"
    meat_intake_radio = browser.find_element(By.XPATH, meat_intake_xpath)
    browser.execute_script("return arguments[0].click()", meat_intake_radio)

    next_page()

    # NU02: Beverages
    bottled_water_input = browser.find_element(By.ID, '95')
    bottled_water_input.send_keys(entry['NU02_01'])

    juice_input = browser.find_element(By.ID, '96')
    juice_input.send_keys(entry['NU02_02'])

    coke_input = browser.find_element(By.ID, '97')
    coke_input.send_keys(entry['NU02_03'])

    milk_input = browser.find_element(By.ID, '98')
    milk_input.send_keys(entry['NU02_04'])

    soymilk_input = browser.find_element(By.ID, '99')
    soymilk_input.send_keys(entry['NU02_05'])

    coffee_input = browser.find_element(By.ID, '100')
    coffee_input.send_keys(entry['NU02_06'])

    tea_input = browser.find_element(By.ID, '101')
    tea_input.send_keys(entry['NU02_07'])

    beer_input = browser.find_element(By.ID, '102')
    beer_input.send_keys(entry['NU02_08'])

    wine_input = browser.find_element(By.ID, '103')
    wine_input.send_keys(entry['NU02_09'])

    tap_water_input = browser.find_element(By.ID, '104')
    tap_water_input.send_keys(entry['NU02_10'])

    next_page()

    # NU03: Food waste frequency
    foodwaste_freq_xpath = f"//input[@name=\"Q13\"][@id=\"{int(entry['NU03'])}\"]"
    foodwaste_freq_radio = browser.find_element(By.XPATH, foodwaste_freq_xpath)
    browser.execute_script("return arguments[0].click()", foodwaste_freq_radio)

    next_page()
    wait.until(ec.visibility_of_element_located((By.XPATH, '//*[@id="resultArea"]/div[3]/h2/strong')))
    run_survey.result_nutrition = browser.find_element(By.XPATH, '//*[@id="resultArea"]/div[3]/h2/strong').text
    print("Nutrition:", run_survey.result_nutrition, "tons")
    next_page()

    #---------------------------------
    # SECTION 4: Leisure
    #---------------------------------

    # LE01: Hobbies
    jogging_input = browser.find_element(By.ID, '111')
    jogging_input.send_keys(entry['LE01_01'])

    biking_input = browser.find_element(By.ID, '112')
    biking_input.send_keys(entry['LE01_02'])

    skating_input = browser.find_element(By.ID, '113')
    skating_input.send_keys(entry['LE01_03'])

    motorcycle_input = browser.find_element(By.ID, '114')
    motorcycle_input.send_keys(entry['LE01_04'])

    music_input = browser.find_element(By.ID, '115')
    music_input.send_keys(entry['LE01_05'])

    next_page()

    # LE02: Facilities
    pool_input = browser.find_element(By.ID, '117')
    pool_input.send_keys(entry['LE02_01'])

    fitness_studio_input = browser.find_element(By.ID, '118')
    fitness_studio_input.send_keys(entry['LE02_02'])

    golf_input = browser.find_element(By.ID, '119')
    golf_input.send_keys(entry['LE02_03'])

    amusement_park_input = browser.find_element(By.ID, '120')
    amusement_park_input.send_keys(entry['LE02_04'])

    cinema_input = browser.find_element(By.ID, '121')
    cinema_input.send_keys(entry['LE02_05'])

    museum_input = browser.find_element(By.ID, '122')
    museum_input.send_keys(entry['LE02_06'])

    tennis_court_input = browser.find_element(By.ID, '123')
    tennis_court_input.send_keys(entry['LE02_07'])

    football_pitch_input = browser.find_element(By.ID, '124')
    football_pitch_input.send_keys(entry['LE02_08'])

    gymnasium_input = browser.find_element(By.ID, '125')
    gymnasium_input.send_keys(entry['LE02_09'])

    climbing_hall_water_input = browser.find_element(By.ID, '126')
    climbing_hall_water_input.send_keys(entry['LE02_10'])

    next_page()

    # LE03: Winter sports
    skiing_xpath = f"//input[@name=\"Q16\"][@id=\"{int(entry['LE03'])}\"]"
    skiing_radio = browser.find_element(By.XPATH, skiing_xpath)
    browser.execute_script("return arguments[0].click()", skiing_radio)

    time.sleep(0.5)

    if entry['LE03_02'] > 0:
        skiing_input = browser.find_element(By.NAME, 'optValQ16')
        skiing_input.send_keys(entry['LE03_02'])

    next_page()
    wait.until(ec.visibility_of_element_located((By.XPATH, '//*[@id="resultArea"]/div[3]/h2/strong')))
    run_survey.result_leisure = browser.find_element(By.XPATH, '//*[@id="resultArea"]/div[3]/h2/strong').text
    print("Leisure:", run_survey.result_leisure, "tons")
    next_page()

    #---------------------------------
    # SECTION 5: Mobility
    #---------------------------------

    # MO01: daily commutes
    daily_car_input = browser.find_element(By.ID, '131')
    daily_car_input.send_keys(entry['MO01_01'])

    daily_scooter_input = browser.find_element(By.ID, '132')
    daily_scooter_input.send_keys(entry['MO01_02'])

    daily_long_dist_train_input = browser.find_element(By.ID, '133')
    daily_long_dist_train_input.send_keys(entry['MO01_03'])

    daily_short_dist_train_input = browser.find_element(By.ID, '134')
    daily_short_dist_train_input.send_keys(entry['MO01_04'])

    daily_bus_input = browser.find_element(By.ID, '135')
    daily_bus_input.send_keys(entry['MO01_05'])

    daily_bike_input = browser.find_element(By.ID, '136')
    daily_bike_input.send_keys(entry['MO01_06'])

    daily_foot_input = browser.find_element(By.ID, '137')
    daily_foot_input.send_keys(entry['MO01_07'])

    next_page()

    # MO02: trips
    trip_car_input = browser.find_element(By.ID, '139')
    trip_car_input.send_keys(entry['MO02_01'])

    trip_scooter_input = browser.find_element(By.ID, '140')
    trip_scooter_input.send_keys(entry['MO02_02'])

    trip_long_dist_train_input = browser.find_element(By.ID, '141')
    trip_long_dist_train_input.send_keys(entry['MO02_03'])

    trip_short_dist_train_input = browser.find_element(By.ID, '142')
    trip_short_dist_train_input.send_keys(entry['MO02_04'])

    trip_bus_input = browser.find_element(By.ID, '143')
    trip_bus_input.send_keys(entry['MO02_05'])

    trip_bike_input = browser.find_element(By.ID, '144')
    trip_bike_input.send_keys(entry['MO02_06'])

    next_page()

    # MO03: Passenger car
    number_cars_input = browser.find_element(By.ID, '146')
    number_cars_input.send_keys(entry['MO03_01'])
    # number drivers defaults to 1 in Wuppetal survey
    number_drivers_input = browser.find_element(By.ID, '147')
    number_drivers_input.send_keys(entry['MO03_02'])

    # car age defaults to 1 in Wuppertal survey
    car_age_input = browser.find_element(By.ID, '148')
    car_age_input.send_keys(entry['MO03_03'])

    next_page()

    # MO04: Vehicle Type
    vehicle_type_xpath = f"//input[@name=\"Q20\"][@id=\"{int(entry['MO04'])}\"]"
    vehicle_type_radio = browser.find_element(By.XPATH, vehicle_type_xpath)
    browser.execute_script("return arguments[0].click()", vehicle_type_radio)

    # if individual drives a car, do car questions
    if entry['MO04'] != 158:
        next_page()

        # MO05: Fuel type
        fuel_type_xpath = f"//input[@name=\"Q21\"][@id=\"{int(entry['MO05'])}\"]"
        fuel_type_radio = browser.find_element(By.XPATH, fuel_type_xpath)
        browser.execute_script("return arguments[0].click()", fuel_type_radio)

        next_page()

        # MO06: Fuel usage
        fuel_usage_xpath = f"//input[@name=\"Q22\"][@id=\"{int(entry['MO06'])}\"]"
        fuel_usage_radio = browser.find_element(By.XPATH, fuel_usage_xpath)
        browser.execute_script("return arguments[0].click()", fuel_usage_radio)

        time.sleep(0.5)

        if entry['MO06_01'] > 0:
            fuel_usage_input = browser.find_element(By.NAME, 'optValQ22')
            fuel_usage_input.send_keys(entry['MO06_01'])

        next_page()

        # MO07: Passengers
        passengers_xpath = f"//input[@name=\"Q23\"][@id=\"{int(entry['MO07'])}\"]"
        passengers_radio = browser.find_element(By.XPATH, passengers_xpath)
        browser.execute_script("return arguments[0].click()", passengers_radio)

        next_page()

    else:
        # skips to next section
        next_page()
        wait.until(ec.visibility_of_element_located(((By.CSS_SELECTOR, 'div[id="skipTrue"]'))))
        browser.find_element(By.CSS_SELECTOR, 'div[id="skipTrue"]').click()


    wait.until(ec.visibility_of_element_located((By.XPATH, '//*[@id="resultArea"]/div[3]/h2/strong')))
    run_survey.result_mobility = browser.find_element(By.XPATH, '//*[@id="resultArea"]/div[3]/h2/strong').text
    print("Mobility:", run_survey.result_mobility, "tons")

    next_page()

    #---------------------------------
    # SECTION 6: Vacation
    #---------------------------------

    # VA01: Travel
    vacation_car_input = browser.find_element(By.ID, '176')
    vacation_car_input.send_keys(entry['VA01_01'])

    vacation_train_input = browser.find_element(By.ID, '177')
    vacation_train_input.send_keys(entry['VA01_02'])

    vacation_bus_input = browser.find_element(By.ID, '178')
    vacation_bus_input.send_keys(entry['VA01_03'])

    vacation_plane_input = browser.find_element(By.ID, '179')
    vacation_plane_input.send_keys(entry['VA01_04'])

    vacation_boat_input = browser.find_element(By.ID, '180')
    vacation_boat_input.send_keys(entry['VA01_05'])

    next_page()

    # VA01: Accomodation
    friend_input = browser.find_element(By.ID, '182')
    friend_input.send_keys(entry['VA02_01'])

    hotel_normal_input = browser.find_element(By.ID, '183')
    hotel_normal_input.send_keys(entry['VA02_02'])

    hotel_luxury_input = browser.find_element(By.ID, '184')
    hotel_luxury_input.send_keys(entry['VA02_03'])

    holiday_house_input = browser.find_element(By.ID, '185')
    holiday_house_input.send_keys(entry['VA02_04'])

    camper_input = browser.find_element(By.ID, '186')
    camper_input.send_keys(entry['VA02_05'])

    tent_input = browser.find_element(By.ID, '187')
    tent_input.send_keys(entry['VA02_06'])

    house_boat_input = browser.find_element(By.ID, '188')
    house_boat_input.send_keys(entry['VA02_07'])

    cruise_input = browser.find_element(By.ID, '189')
    cruise_input.send_keys(entry['VA02_08'])

    next_page()
    wait.until(ec.visibility_of_element_located((By.XPATH, '//*[@id="resultArea"]/div[3]/h2/strong')))
    run_survey.result_vacation = browser.find_element(By.XPATH, '//*[@id="resultArea"]/div[3]/h2/strong').text
    print("Vacation:", run_survey.result_vacation, "tons")

    # overall results
    browser.find_element(By.CSS_SELECTOR, 'p[id="next-button-result"]').click()

    # skips sociodemographic data
    wait.until(ec.visibility_of_element_located((By.CSS_SELECTOR, 'div[id="skipSocialData"]')))
    browser.find_element(By.CSS_SELECTOR, 'div[id="skipSocialData"]').click()

    # gets total result
    wait.until(ec.visibility_of_element_located((By.XPATH, '//*[@id="resultArea"]/div[5]/h1/strong')))
    run_survey.result_total = browser.find_element(By.XPATH, '//*[@id="resultArea"]/div[5]/h1/strong').text
    print("Total footprint", run_survey.result_total, "tons")
    print(" ")

'''
RUN SECTION
iterates the process, only taking cases that are not already in the output workbook
This is necessary because the program encounters timeout error since it would take
~2 hours to run from start to finish.

We run multiple instances of the program and append the output workbook.
It's a necessary, if inelegant, solution.
'''

for i in range(91, len(dataframe) + 1):
    entry = dataframe.iloc[i]
    run_survey()

    # fill EI numbers to workbook
    sheet.cell(row=i+2, column=1).value=entry['CASE']
    sheet.cell(row=i+2, column=2).value=float(run_survey.result_housing)
    sheet.cell(row=i+2, column=3).value=float(run_survey.result_consumer)
    sheet.cell(row=i+2, column=4).value=float(run_survey.result_nutrition)
    sheet.cell(row=i+2, column=5).value=float(run_survey.result_leisure)
    sheet.cell(row=i+2, column=6).value=float(run_survey.result_mobility)
    sheet.cell(row=i+2, column=7).value=float(run_survey.result_vacation)
    sheet.cell(row=i+2, column=8).value=float(run_survey.result_total)
    # save over old file
    wb.save('impact_results.xlsx')

    # restart survey
    browser.get(url)
